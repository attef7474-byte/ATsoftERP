import openpyxl, json, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook('docs/data-prep/batch-e/ATsoft_ERP_source.xlsx', data_only=True)
sheets = []
for name in wb.sheetnames:
    ws = wb[name]
    rows = ws.max_row or 0
    cols = ws.max_column or 0
    headers = []
    if rows > 0:
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value)[:50])
    sheets.append({'name': name, 'rows': rows, 'cols': cols, 'headers': headers[:10]})
wb.close()
with open('docs/data-prep/batch-e/sheet_inventory.json', 'w', encoding='utf-8') as f:
    json.dump({'total_sheets': len(sheets), 'sheets': sheets}, f, ensure_ascii=False, indent=2)
print('Inventory written: ' + str(len(sheets)) + ' sheets')
for s in sheets:
    h = ', '.join(s['headers'][:5])
    print('  ' + s['name'] + ': ' + str(s['rows']) + 'r x ' + str(s['cols']) + 'c | ' + h)
