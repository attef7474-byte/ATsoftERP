import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

FILE_PATH = r'C:\Users\attef\PycharmProjects\Trae\ATsofterp\docs\data-prep\batch-e\ATsoft_ERP_source.xlsx'
OUTPUT_PATH = r'C:\Users\attef\PycharmProjects\Trae\ATsofterp\docs\data-prep\batch-e\core_sheets_data.json'

SHEETS = [
    '03_الشركات',
    '04_الفروع',
    '05_الإدارات',
    '06_الأقسام',
    '36_المسميات_الوظيفية',
    '37_الأشخاص_التشغيليون',
    '38_تعيينات_الأشخاص',
    '39_التسلسل_الإداري',
    '24_كادر_الصيانة',
    '40_مسؤوليات_الصيانة_v2',
]

wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)

result = {}

for sheet_name in SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"WARNING: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result[sheet_name] = {"headers": [], "rows": []}
        continue
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data_rows = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if isinstance(val, float) and val == int(val):
                val = int(val)
            row_dict[h] = val
        data_rows.append(row_dict)
    result[sheet_name] = {"headers": headers, "rows": data_rows}
    print(f"{sheet_name}: {len(data_rows)} rows, {len(headers)} columns")

wb.close()

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\nWritten to {OUTPUT_PATH}")
